"""Offline tests for the wb_pinksheet connector — no network, recorded fixtures only.

Fixtures (see fixtures/README.md): real excerpts recorded from the World Bank
commodity-markets landing page and CMO-Historical-Data-Monthly.xlsx on
2026-07-18, trimmed to header rows + first 6 + last 60 data rows.
"""

import io
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from erda_contracts.errors import ContractViolation, SourceUnavailable
from erda_contracts.ledger import read_ledger
from erda_ingestion import wb_pinksheet
from erda_ingestion.base import FetchResult, run_connector

FIXTURES = Path(__file__).parent / "fixtures"
XLSX_FIXTURE = FIXTURES / "wb_pinksheet_monthly_sample.xlsx"
HTML_FIXTURE = FIXTURES / "wb_pinksheet_landing_sample.html"


def _xlsx_bytes() -> bytes:
    return XLSX_FIXTURE.read_bytes()


def _workbook_bytes(rows: list[list], sheet: str = wb_pinksheet.SHEET) -> bytes:
    """Build in-memory XLSX bytes for layout-drift failure paths (synthetic)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --- href scraper -----------------------------------------------------------


def test_find_monthly_xlsx_url_picks_monthly_not_annual():
    url = wb_pinksheet.find_monthly_xlsx_url(HTML_FIXTURE.read_text())
    assert url == (
        "https://thedocs.worldbank.org/en/doc/"
        "74e8be41ceb20fa0da750cda2f6b9e4e-0050012026/related/CMO-Historical-Data-Monthly.xlsx"
    )


def test_find_monthly_xlsx_url_resolves_relative_href():
    html = '<a href="/en/doc/abc123/related/CMO-Historical-Data-Monthly.xlsx">Monthly</a>'
    url = wb_pinksheet.find_monthly_xlsx_url(html)
    assert url == (
        "https://www.worldbank.org/en/doc/abc123/related/CMO-Historical-Data-Monthly.xlsx"
    )


def test_find_monthly_xlsx_url_missing_raises():
    with pytest.raises(SourceUnavailable, match="rotating-path scrape failed"):
        wb_pinksheet.find_monthly_xlsx_url("<html><body><p>no links here</p></body></html>")


# --- normalize --------------------------------------------------------------


def test_normalize_long_format_units_and_sentinels():
    df = wb_pinksheet.normalize(_xlsx_bytes())
    assert list(df.columns) == ["month", "commodity", "price", "units"]
    assert pd.api.types.is_datetime64_any_dtype(df["month"])
    assert set(df["commodity"]) == {"crude_avg", "brent", "dubai", "wti", "coal_au"}

    # units come from row 6, not assumption: crude in usd_bbl, coal in usd_mt
    assert set(df.loc[df["commodity"] != "coal_au", "units"]) == {"usd_bbl"}
    assert set(df.loc[df["commodity"] == "coal_au", "units"]) == {"usd_mt"}

    # "1960M01" date layout parses to the first of the month
    jan60 = df[df["month"] == pd.Timestamp("1960-01-01")]
    assert jan60.set_index("commodity")["price"]["crude_avg"] == 1.6
    # "…" sentinel (WTI and coal not quoted in 1960M01) dropped, never imputed
    assert set(jan60["commodity"]) == {"crude_avg", "brent", "dubai"}

    # spot-check the latest recorded month against the real sheet (2026M06)
    jun26 = df[df["month"] == pd.Timestamp("2026-06-01")].set_index("commodity")["price"]
    assert jun26["brent"] == 85.4
    assert jun26["coal_au"] == 138.5

    assert not df.duplicated(["month", "commodity"]).any()


def test_normalize_missing_sheet_raises():
    content = _workbook_bytes([["not the pink sheet"]], sheet="Wrong Sheet")
    with pytest.raises(SourceUnavailable, match="Monthly Prices"):
        wb_pinksheet.normalize(content)


def test_normalize_renamed_header_raises():
    """Row-5 name drift (e.g. Brent renamed) must fail loudly, not drop the series."""
    wb = openpyxl.load_workbook(io.BytesIO(_xlsx_bytes()))
    ws = wb[wb_pinksheet.SHEET]
    ws.cell(row=5, column=3).value = "Crude oil, Brent (new basis)"
    buf = io.BytesIO()
    wb.save(buf)
    with pytest.raises(SourceUnavailable, match="Crude oil, Brent"):
        wb_pinksheet.normalize(buf.getvalue())


def test_normalize_unit_drift_raises():
    """A tracked column whose row-6 unit label changed must not be mislabelled."""
    wb = openpyxl.load_workbook(io.BytesIO(_xlsx_bytes()))
    ws = wb[wb_pinksheet.SHEET]
    ws.cell(row=6, column=2).value = "($/toe)"
    buf = io.BytesIO()
    wb.save(buf)
    with pytest.raises(SourceUnavailable, match=r"\$\/toe"):
        wb_pinksheet.normalize(buf.getvalue())


def test_normalize_bad_date_layout_raises():
    rows = [
        ["title"], ["subtitle"], ["note"], ["updated"],
        [None, "Crude oil, average", "Crude oil, Brent", "Crude oil, Dubai",
         "Crude oil, WTI", "Coal, Australian"],
        [None, "($/bbl)", "($/bbl)", "($/bbl)", "($/bbl)", "($/mt)"],
        ["Jan-1960", 1.6, 1.6, 1.6, "…", "…"],  # date layout drifted
    ]
    with pytest.raises(SourceUnavailable, match="YYYYMmm"):
        wb_pinksheet.normalize(_workbook_bytes(rows))


# --- runner integration (fake fetch, tmp_path) ------------------------------


def test_runner_full_path_writes_provenance_and_ledger(tmp_path):
    frame = wb_pinksheet.normalize(_xlsx_bytes())
    xlsx_url = "https://thedocs.worldbank.org/en/doc/hash/related/CMO-Historical-Data-Monthly.xlsx"

    def fake_fetch() -> FetchResult:
        return FetchResult(frame=frame, source_url=xlsx_url)

    entry = run_connector(
        source_id=wb_pinksheet.SOURCE_ID,
        transform_version=wb_pinksheet.TRANSFORM_VERSION,
        schema=wb_pinksheet.SCHEMA,
        fetch=fake_fetch,
        table=wb_pinksheet.TABLE,
        root=tmp_path,
    )
    assert entry.rows == len(frame)
    written = pd.read_parquet(tmp_path / "wb_pinksheet_prices.parquet")
    # every persisted number carries provenance (§0 rule 5)
    for col in ["source_id", "retrieved_at", "source_url", "transform_version"]:
        assert col in written.columns
    assert (written["source_id"] == "wb_pinksheet").all()
    assert (written["source_url"] == xlsx_url).all()
    assert read_ledger(tmp_path)[0].table == "wb_pinksheet_prices"


def test_runner_rejects_contract_violation(tmp_path):
    bad = wb_pinksheet.normalize(_xlsx_bytes())
    bad.loc[0, "price"] = -1.0  # negative price violates the contract

    def fake_fetch() -> FetchResult:
        return FetchResult(frame=bad, source_url="https://example.invalid/pink.xlsx")

    with pytest.raises(ContractViolation):
        run_connector(
            source_id=wb_pinksheet.SOURCE_ID,
            transform_version=wb_pinksheet.TRANSFORM_VERSION,
            schema=wb_pinksheet.SCHEMA,
            fetch=fake_fetch,
            table=wb_pinksheet.TABLE,
            root=tmp_path,
        )
    # nothing bad persisted, ledger untouched
    assert read_ledger(tmp_path) == []


def test_runner_rejects_non_numeric_junk_as_nan(tmp_path):
    """A value that is neither numeric nor '…' must surface as ContractViolation."""
    wb = openpyxl.load_workbook(io.BytesIO(_xlsx_bytes()))
    ws = wb[wb_pinksheet.SHEET]
    ws.cell(row=8, column=2).value = "n.a."  # undocumented sentinel
    buf = io.BytesIO()
    wb.save(buf)
    frame = wb_pinksheet.normalize(buf.getvalue())
    assert frame["price"].isna().sum() == 1  # kept as NaN, not silently dropped

    def fake_fetch() -> FetchResult:
        return FetchResult(frame=frame, source_url="https://example.invalid/pink.xlsx")

    with pytest.raises(ContractViolation):
        run_connector(
            source_id=wb_pinksheet.SOURCE_ID,
            transform_version=wb_pinksheet.TRANSFORM_VERSION,
            schema=wb_pinksheet.SCHEMA,
            fetch=fake_fetch,
            table=wb_pinksheet.TABLE,
            root=tmp_path,
        )
    assert read_ledger(tmp_path) == []
