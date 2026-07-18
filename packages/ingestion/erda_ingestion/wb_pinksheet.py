"""wb_pinksheet — World Bank Commodity Markets "Pink Sheet" monthly prices.

Cadence: monthly (spec §4; registry SLA 40 days). No key.

Registry drift notes (live-verified 2026-07-18):
- CMO-Historical-Data-Monthly.xlsx is served from thedocs.worldbank.org under a
  ROTATING hashed path — every run scrapes the landing page
  (https://www.worldbank.org/en/research/commodity-markets) for the current
  href. Never hardcode the download URL.
- "Monthly Prices" sheet layout: commodity names on row 5, units on row 6
  (e.g. "($/bbl)", "($/mt)"), data from row 7. Dates formatted "1960M01".
  Missing values are the ellipsis sentinel "…" — dropped, never imputed.
  Verified through 2026M06.

Long format: one row per (month, commodity). Crude benchmarks in usd_bbl plus
Australian coal (usd_mt) as a cross-check series. Units are read from row 6,
not assumed — an unrecognized unit label for a tracked column means the sheet
layout drifted and the connector raises instead of guessing.
"""

from __future__ import annotations

import io
import re
from urllib.parse import urljoin

import openpyxl
import pandas as pd
import pandera.pandas as pa
from bs4 import BeautifulSoup

from erda_contracts.errors import SourceUnavailable
from erda_ingestion.base import FetchResult, http_get

SOURCE_ID = "wb_pinksheet"
TRANSFORM_VERSION = "wb_pinksheet:1.0.0"
TABLE = "wb_pinksheet_prices"

LANDING_URL = "https://www.worldbank.org/en/research/commodity-markets"
XLSX_NAME = "CMO-Historical-Data-Monthly.xlsx"
SHEET = "Monthly Prices"

#: Row-5 header labels (exact, post-strip) → tidy commodity ids.
COMMODITIES = {
    "Crude oil, average": "crude_avg",
    "Crude oil, Brent": "brent",
    "Crude oil, Dubai": "dubai",
    "Crude oil, WTI": "wti",
    "Coal, Australian": "coal_au",
}

#: Row-6 unit labels (exact, post-strip) → unit codes.
UNIT_LABELS = {
    "($/bbl)": "usd_bbl",
    "($/mt)": "usd_mt",
}

MISSING_SENTINEL = "…"  # "…" per registry
_DATE_RE = re.compile(r"^\d{4}M\d{2}$")  # "1960M01"

SCHEMA = pa.DataFrameSchema(
    {
        "month": pa.Column(pa.DateTime, nullable=False),
        "commodity": pa.Column(str, pa.Check.isin(list(COMMODITIES.values())), nullable=False),
        "price": pa.Column(float, pa.Check.gt(0), nullable=False),
        "units": pa.Column(str, pa.Check.isin(list(UNIT_LABELS.values())), nullable=False),
    },
    unique=["month", "commodity"],
)


def find_monthly_xlsx_url(html: str) -> str:
    """Scrape the landing page for the current (rotating) Monthly XLSX href."""
    soup = BeautifulSoup(html, "html.parser")
    for anchor in soup.find_all("a", href=True):
        href = anchor["href"]
        if XLSX_NAME in href:
            return urljoin(LANDING_URL, href)
    raise SourceUnavailable(
        SOURCE_ID, f"no {XLSX_NAME} link on {LANDING_URL} — rotating-path scrape failed"
    )


def normalize(content: bytes) -> pd.DataFrame:
    """Pink Sheet XLSX bytes → tidy long frame [month, commodity, price, units].

    Layout contract (registry, verified 2026-07-18): names row 5, units row 6,
    data from row 7, "1960M01" dates, "…" missing sentinel. Tracked columns are
    located by row-5 name, not position. Layout drift (missing sheet, missing
    name, unknown unit, unparseable date) raises SourceUnavailable; a value
    that is neither numeric nor the documented sentinel becomes NaN and is
    rejected by the contract downstream — never silently dropped.
    """
    workbook = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    if SHEET not in workbook.sheetnames:
        raise SourceUnavailable(SOURCE_ID, f"sheet {SHEET!r} not in workbook — layout drifted")
    rows = list(workbook[SHEET].iter_rows(values_only=True))
    if len(rows) < 7:
        raise SourceUnavailable(SOURCE_ID, f"sheet {SHEET!r} has no data rows — layout drifted")
    names, units = rows[4], rows[5]  # registry: names row 5, units row 6

    columns: dict[int, tuple[str, str]] = {}  # col index → (commodity, unit code)
    for idx, name in enumerate(names):
        label = name.strip() if isinstance(name, str) else None
        if label not in COMMODITIES:
            continue
        # read_only mode may trim trailing cells, so guard the row width
        unit_label = units[idx] if idx < len(units) else None
        if isinstance(unit_label, str):
            unit_label = unit_label.strip()
        if unit_label not in UNIT_LABELS:
            raise SourceUnavailable(
                SOURCE_ID, f"unit for {label!r} is {unit_label!r}, not in {list(UNIT_LABELS)}"
            )
        columns[idx] = (COMMODITIES[label], UNIT_LABELS[unit_label])
    missing = sorted(set(COMMODITIES) - {names[i].strip() for i in columns})
    if missing:
        raise SourceUnavailable(SOURCE_ID, f"row-5 names missing {missing} — layout drifted")

    records: list[tuple[str, str, object, str]] = []
    for row in rows[6:]:  # registry: data from row 7
        if all(cell is None for cell in row):
            continue  # trailing blank row, not data
        date_label = row[0]
        if not (isinstance(date_label, str) and _DATE_RE.match(date_label.strip())):
            raise SourceUnavailable(
                SOURCE_ID, f"date cell {date_label!r} does not match 'YYYYMmm' layout"
            )
        for idx, (commodity, unit) in columns.items():
            value = row[idx] if idx < len(row) else None
            if value is None or (isinstance(value, str) and value.strip() == MISSING_SENTINEL):
                continue  # documented missing sentinel — dropped, never imputed
            records.append((date_label.strip(), commodity, value, unit))
    if not records:
        raise SourceUnavailable(SOURCE_ID, "no data rows parsed — layout drifted")

    df = pd.DataFrame(records, columns=["month", "commodity", "price", "units"])
    df["month"] = pd.to_datetime(df["month"], format="%YM%m")
    # Non-numeric junk (anything besides the documented "…") becomes NaN here
    # and fails the non-nullable contract — loud, not silent (§0 rule 4).
    df["price"] = pd.to_numeric(df["price"], errors="coerce")
    return df


def fetch() -> FetchResult:
    landing = http_get(LANDING_URL, SOURCE_ID)
    xlsx_url = find_monthly_xlsx_url(landing.text)
    resp = http_get(xlsx_url, SOURCE_ID)
    # Provenance records the resolved (hashed, rotating) URL actually fetched.
    return FetchResult(frame=normalize(resp.content), source_url=xlsx_url)
